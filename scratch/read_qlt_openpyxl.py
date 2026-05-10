import openpyxl

file_path = r'd:\CC\SQA\Code BE-TEST\TestDoAn1\4 - Unit test.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'Module QLT' in wb.sheetnames:
        sheet = wb['Module QLT']
        print(f"--- Data from sheet: Module QLT ---\n")
        
        # Đọc 50 dòng đầu tiên để phân tích
        for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
            print(row)
    else:
        print(f"Sheet 'Module QLT' not found. Available sheets: {wb.sheetnames}")

except Exception as e:
    print(f"Error reading Excel with openpyxl: {e}")
